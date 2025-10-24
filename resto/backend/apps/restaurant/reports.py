"""
Report API ViewSet for analytics and reporting functionality
Provides endpoints for sales, expenses, product performance, and trend analysis
"""

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Avg, Count, F, Q, DecimalField, OuterRef, Subquery
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, datetime
from decimal import Decimal
from io import BytesIO

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Order, OrderItem, Payment, Product, InventoryTransaction, CashierSession
from .permissions import IsManagerOrAdmin


class ReportViewSet(viewsets.ViewSet):
    """
    ViewSet for generating various reports with time-based filtering

    Supports period filters: today, week, month, year, custom
    """
    permission_classes = [IsAuthenticated]

    def get_order_total_sum(self):
        """
        Returns a Sum expression for calculating order totals from items
        Use this in aggregate() calls directly to avoid nested aggregation issues
        """
        return Coalesce(
            Sum(
                F('items__unit_price') * F('items__quantity') - F('items__discount_amount'),
                output_field=DecimalField()
            ),
            Decimal('0')
        )

    def get_order_total_avg(self):
        """
        Returns an Avg expression for calculating average order value from items
        """
        return Coalesce(
            Avg(
                F('items__unit_price') * F('items__quantity') - F('items__discount_amount'),
                output_field=DecimalField()
            ),
            Decimal('0')
        )

    def get_date_range(self, period, start_date=None, end_date=None):
        """
        Helper to calculate date ranges based on period

        Args:
            period: 'today', 'week', 'month', 'year', or 'custom'
            start_date: Start date for custom period (YYYY-MM-DD)
            end_date: End date for custom period (YYYY-MM-DD)

        Returns:
            tuple: (start_date, end_date)
        """
        today = timezone.now().date()

        if period == 'today':
            return today, today
        elif period == 'week':
            # Start from Monday of current week
            start = today - timedelta(days=today.weekday())
            return start, today
        elif period == 'month':
            # Start from first day of current month
            start = today.replace(day=1)
            return start, today
        elif period == 'year':
            # Start from January 1st of current year
            start = today.replace(month=1, day=1)
            return start, today
        elif period == 'custom' and start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                return start, end
            except ValueError:
                return today, today

        # Default to today
        return today, today

    def get_previous_period_range(self, start_date, end_date):
        """
        Calculate the previous period of the same length

        Args:
            start_date: Start date of current period
            end_date: End date of current period

        Returns:
            tuple: (prev_start_date, prev_end_date)
        """
        period_length = (end_date - start_date).days + 1
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_length - 1)
        return prev_start, prev_end

    @action(detail=False, methods=['get'])
    def sales(self, request):
        """
        Sales analytics endpoint

        Query params:
            - period: today, week, month, year, custom
            - start_date: YYYY-MM-DD (for custom period)
            - end_date: YYYY-MM-DD (for custom period)
            - branch: branch ID (optional)

        Returns:
            - summary: Total revenue, orders, avg order value, expenses, profit
            - daily_breakdown: Day-by-day sales data
            - comparison: Comparison with previous period
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)
        prev_start, prev_end = self.get_previous_period_range(start_date, end_date)

        # Base query for completed orders
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            status='COMPLETED'
        )

        prev_orders = Order.objects.filter(
            created_at__date__gte=prev_start,
            created_at__date__lte=prev_end,
            status='COMPLETED'
        )

        if branch_id and branch_id != 'all':
            orders = orders.filter(branch_id=branch_id)
            prev_orders = prev_orders.filter(branch_id=branch_id)

        # Annotate with calculated totals

        # Calculate summary statistics
        summary = orders.aggregate(
            total_revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            total_orders=Count('id'),
            avg_order_value=Coalesce(Avg(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0'))
        )

        prev_summary = prev_orders.aggregate(
            total_revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            total_orders=Count('id')
        )

        # Calculate expenses (inventory movements with type 'OUT' or 'ADJUSTMENT')
        expenses = InventoryTransaction.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            transaction_type__in=['IN', 'ADJUST']
        )

        if branch_id and branch_id != 'all':
            expenses = expenses.filter(inventory__branch_id=branch_id)

        total_expenses = expenses.aggregate(
            total=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=DecimalField()), Decimal('0'))
        )['total']

        # Calculate net profit
        net_profit = summary['total_revenue'] - total_expenses
        profit_margin = (net_profit / summary['total_revenue'] * 100) if summary['total_revenue'] > 0 else 0

        # Daily breakdown
        daily_data = orders.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            orders_count=Count('id'),
            avg_order_value=Coalesce(Avg(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0'))
        ).order_by('date')

        # Get top product for each day
        daily_breakdown = []
        for day in daily_data:
            # Find top product for this day
            top_product = OrderItem.objects.filter(
                order__created_at__date=day['date'],
                order__status='COMPLETED'
            ).values('product__name').annotate(
                total_qty=Sum('quantity')
            ).order_by('-total_qty').first()

            daily_breakdown.append({
                'date': day['date'].strftime('%Y-%m-%d'),
                'revenue': float(day['revenue']),
                'orders': day['orders_count'],
                'avg_order_value': float(day['avg_order_value']),
                'top_product': top_product['product__name'] if top_product else 'N/A'
            })

        # Calculate growth metrics
        revenue_growth = 0
        orders_growth = 0
        if prev_summary['total_revenue'] > 0:
            revenue_growth = ((summary['total_revenue'] - prev_summary['total_revenue']) / prev_summary['total_revenue'] * 100)
        if prev_summary['total_orders'] > 0:
            orders_growth = ((summary['total_orders'] - prev_summary['total_orders']) / prev_summary['total_orders'] * 100)

        return Response({
            'summary': {
                'total_revenue': float(summary['total_revenue']),
                'total_orders': summary['total_orders'],
                'avg_order_value': float(summary['avg_order_value']),
                'total_expenses': float(total_expenses),
                'net_profit': float(net_profit),
                'profit_margin': float(profit_margin),
            },
            'daily_breakdown': daily_breakdown,
            'comparison': {
                'current_period': {
                    'revenue': float(summary['total_revenue']),
                    'orders': summary['total_orders'],
                },
                'previous_period': {
                    'revenue': float(prev_summary['total_revenue']),
                    'orders': prev_summary['total_orders'],
                },
                'growth': {
                    'revenue_percent': float(revenue_growth),
                    'orders_percent': float(orders_growth),
                }
            }
        })

    @action(detail=False, methods=['get'])
    def expenses(self, request):
        """
        Expense analytics endpoint

        Query params:
            - period: today, week, month, year, custom
            - start_date, end_date, branch

        Returns:
            - total_expenses: Sum of all expenses
            - breakdown: Expenses by category with trends
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)
        prev_start, prev_end = self.get_previous_period_range(start_date, end_date)

        # Current period expenses
        expenses_query = InventoryTransaction.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            transaction_type__in=['IN', 'ADJUST']
        )

        # Previous period expenses
        prev_expenses_query = InventoryTransaction.objects.filter(
            created_at__date__gte=prev_start,
            created_at__date__lte=prev_end,
            transaction_type__in=['IN', 'ADJUST']
        )

        if branch_id and branch_id != 'all':
            expenses_query = expenses_query.filter(inventory__branch_id=branch_id)
            prev_expenses_query = prev_expenses_query.filter(inventory__branch_id=branch_id)

        # Calculate total expenses
        total_expenses = expenses_query.aggregate(
            total=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=DecimalField()), Decimal('0'))
        )['total']

        # Breakdown by category (using inventory item location as proxy for category)
        categories = {
            'KITCHEN': 'Bahan Baku',
            'WAREHOUSE': 'Persediaan Gudang',
        }

        breakdown = []
        for location, category_name in categories.items():
            current = expenses_query.filter(inventory__location=location).aggregate(
                amount=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=DecimalField()), Decimal('0'))
            )['amount']

            previous = prev_expenses_query.filter(inventory__location=location).aggregate(
                amount=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=DecimalField()), Decimal('0'))
            )['amount']

            percentage = (current / total_expenses * 100) if total_expenses > 0 else 0

            # Determine trend
            if previous == 0:
                trend = 'stable'
                change_percent = 0
            else:
                change_percent = ((current - previous) / previous * 100)
                if change_percent > 5:
                    trend = 'up'
                elif change_percent < -5:
                    trend = 'down'
                else:
                    trend = 'stable'

            breakdown.append({
                'category': category_name,
                'amount': float(current),
                'percentage': float(percentage),
                'trend': trend,
                'previous_period_amount': float(previous),
                'change_percentage': float(change_percent)
            })

        # Add labor costs (from cashier sessions)
        # This is a placeholder - you may want to add actual payroll tracking
        labor_costs = CashierSession.objects.filter(
            opened_at__date__gte=start_date,
            opened_at__date__lte=end_date,
            status='CLOSED'
        )

        if branch_id and branch_id != 'all':
            labor_costs = labor_costs.filter(branch_id=branch_id)

        labor_count = labor_costs.count()
        # Estimate: assume 100,000 per session as labor cost
        estimated_labor = labor_count * 100000

        if estimated_labor > 0:
            breakdown.append({
                'category': 'Gaji Karyawan',
                'amount': float(estimated_labor),
                'percentage': float((estimated_labor / (total_expenses + estimated_labor) * 100)),
                'trend': 'stable',
                'previous_period_amount': float(estimated_labor),
                'change_percentage': 0
            })

        # Calculate previous period total for comparison
        prev_total = prev_expenses_query.aggregate(
            total=Coalesce(Sum(F('quantity') * F('unit_cost'), output_field=DecimalField()), Decimal('0'))
        )['total']

        # Calculate growth percentage
        if prev_total > 0:
            growth_percentage = ((total_expenses - prev_total) / prev_total * 100)
        else:
            growth_percentage = 0

        return Response({
            'summary': {
                'total_expenses': float(total_expenses),
                'previous_total': float(prev_total)
            },
            'comparison': {
                'growth_percentage': float(growth_percentage)
            },
            'by_category': breakdown
        })

    @action(detail=False, methods=['get'])
    def products(self, request):
        """
        Product performance analytics

        Query params:
            - period, start_date, end_date, branch
            - limit: Number of top products (default: 10)

        Returns:
            - top_products: Best selling products
            - low_performers: Products with low sales
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        limit = int(request.query_params.get('limit', 10))

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)

        # Query order items for completed orders
        order_items = OrderItem.objects.filter(
            order__created_at__date__gte=start_date,
            order__created_at__date__lte=end_date,
            order__status='COMPLETED'
        )

        if branch_id and branch_id != 'all':
            order_items = order_items.filter(order__branch_id=branch_id)

        # Aggregate by product
        product_stats = order_items.values(
            'product_id',
            'product__name'
        ).annotate(
            quantity_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('unit_price'), output_field=DecimalField())
        ).order_by('-quantity_sold')

        # Calculate total revenue for contribution percentages
        total_revenue = sum(float(item['revenue']) for item in product_stats)

        # Top products
        top_products = []
        for item in product_stats[:limit]:
            # Calculate profit (simplified - using 50% margin assumption)
            # In real scenario, use recipe costs
            revenue = float(item['revenue'])
            estimated_profit = revenue * 0.5
            profit_margin = 50.0
            contribution_percentage = (revenue / total_revenue * 100) if total_revenue > 0 else 0

            top_products.append({
                'product_id': item['product_id'],
                'product_name': item['product__name'],
                'quantity_sold': item['quantity_sold'],
                'revenue': revenue,
                'profit': estimated_profit,
                'profit_margin': profit_margin,
                'contribution_percentage': contribution_percentage
            })

        # Low performers (bottom 5)
        low_performers = []
        for item in product_stats[max(0, len(product_stats)-5):]:
            low_performers.append({
                'product_id': item['product_id'],
                'product_name': item['product__name'],
                'quantity_sold': item['quantity_sold'],
                'revenue': float(item['revenue'])
            })

        return Response({
            'top_products': top_products,
            'low_performers': low_performers
        })

    @action(detail=False, methods=['get'])
    def trends(self, request):
        """
        Trend analysis over time

        Query params:
            - period, start_date, end_date, branch
            - metric: 'revenue', 'orders', 'customers' (default: revenue)

        Returns:
            - time_series: Daily/weekly data points
            - comparison: Current vs previous period
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        metric = request.query_params.get('metric', 'revenue')

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)
        prev_start, prev_end = self.get_previous_period_range(start_date, end_date)

        # Query orders
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            status='COMPLETED'
        )

        prev_orders = Order.objects.filter(
            created_at__date__gte=prev_start,
            created_at__date__lte=prev_end,
            status='COMPLETED'
        )

        if branch_id and branch_id != 'all':
            orders = orders.filter(branch_id=branch_id)
            prev_orders = prev_orders.filter(branch_id=branch_id)

        # Annotate with calculated totals

        # Time series data
        time_series = orders.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            orders_count=Count('id'),
            unique_customers=Count('table', distinct=True)  # Using table as proxy for customers
        ).order_by('date')

        time_series_data = [
            {
                'date': item['date'].strftime('%Y-%m-%d'),
                'revenue': float(item['revenue']),
                'orders': item['orders_count'],
                'unique_customers': item['unique_customers']
            }
            for item in time_series
        ]

        # Current period aggregates
        current_agg = orders.aggregate(
            revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            orders=Count('id')
        )

        prev_agg = prev_orders.aggregate(
            revenue=Coalesce(Sum(F('items__unit_price') * F('items__quantity') - F('items__discount_amount'), output_field=DecimalField()), Decimal('0')),
            orders=Count('id')
        )

        # Calculate daily averages
        days_in_period = (end_date - start_date).days + 1
        prev_days = (prev_end - prev_start).days + 1

        current_avg = current_agg['revenue'] / days_in_period if days_in_period > 0 else 0
        prev_avg = prev_agg['revenue'] / prev_days if prev_days > 0 else 0

        # Growth calculations
        revenue_growth = 0
        orders_growth = 0
        if prev_agg['revenue'] > 0:
            revenue_growth = ((current_agg['revenue'] - prev_agg['revenue']) / prev_agg['revenue'] * 100)
        if prev_agg['orders'] > 0:
            orders_growth = ((current_agg['orders'] - prev_agg['orders']) / prev_agg['orders'] * 100)

        return Response({
            'time_series': time_series_data,
            'comparison': {
                'current_period': {
                    'revenue': float(current_agg['revenue']),
                    'orders': current_agg['orders'],
                    'avg_daily': float(current_avg)
                },
                'previous_period': {
                    'revenue': float(prev_agg['revenue']),
                    'orders': prev_agg['orders'],
                    'avg_daily': float(prev_avg)
                },
                'growth': {
                    'revenue_percent': float(revenue_growth),
                    'orders_percent': float(orders_growth)
                }
            }
        })

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """
        Export full report as PDF with formatting

        Query params: period, branch, start_date, end_date

        Returns: PDF file with complete formatted report
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)

        # Fetch all report data
        # Re-use existing action methods to get data
        request._request.GET = request._request.GET.copy()
        sales_data = self.sales(request).data
        expenses_data = self.expenses(request).data
        products_data = self.products(request).data
        trends_data = self.trends(request).data

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#005357'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        elements.append(Paragraph('Laporan Penjualan & Keuangan', title_style))
        elements.append(Paragraph(f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}',
                                  styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary Section
        elements.append(Paragraph('Ringkasan Penjualan', styles['Heading2']))
        summary_data = [
            ['Metrik', 'Nilai'],
            ['Total Pendapatan', f"Rp {sales_data['summary']['total_revenue']:,.0f}"],
            ['Total Pesanan', f"{sales_data['summary']['total_orders']}"],
            ['Rata-rata Nilai Pesanan', f"Rp {sales_data['summary']['avg_order_value']:,.0f}"],
            ['Total Pengeluaran', f"Rp {sales_data['summary']['total_expenses']:,.0f}"],
            ['Laba Bersih', f"Rp {sales_data['summary']['net_profit']:,.0f}"],
            ['Margin Laba', f"{sales_data['summary']['profit_margin']:.1f}%"],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005357')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Expenses Breakdown
        elements.append(Paragraph('Rincian Pengeluaran', styles['Heading2']))
        expense_data = [['Kategori', 'Jumlah', 'Persentase', 'Tren']]
        for item in expenses_data['by_category']:
            expense_data.append([
                item['category'],
                f"Rp {item['amount']:,.0f}",
                f"{item['percentage']:.1f}%",
                item['trend'].upper()
            ])

        expense_table = Table(expense_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005357')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(expense_table)
        elements.append(Spacer(1, 20))

        # Top Products
        elements.append(Paragraph('Produk Terlaris', styles['Heading2']))
        product_data = [['Produk', 'Terjual', 'Pendapatan', 'Kontribusi']]
        for item in products_data['top_products']:
            product_data.append([
                item['product_name'],
                f"{item['quantity_sold']}",
                f"Rp {item['revenue']:,.0f}",
                f"{item['contribution_percentage']:.1f}%"
            ])

        product_table = Table(product_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1*inch])
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005357')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(product_table)
        elements.append(Spacer(1, 20))

        # Trends
        elements.append(Paragraph('Analisis Tren', styles['Heading2']))
        trend_comp = trends_data['comparison']
        trend_data = [
            ['Metrik', 'Periode Saat Ini', 'Periode Sebelumnya', 'Pertumbuhan'],
            ['Pendapatan',
             f"Rp {trend_comp['current_period']['revenue']:,.0f}",
             f"Rp {trend_comp['previous_period']['revenue']:,.0f}",
             f"{trend_comp['growth']['revenue_percent']:.1f}%"],
            ['Jumlah Pesanan',
             f"{trend_comp['current_period']['orders']}",
             f"{trend_comp['previous_period']['orders']}",
             f"{trend_comp['growth']['orders_percent']:.1f}%"],
        ]

        trend_table = Table(trend_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005357')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(trend_table)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="laporan_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf"'
        response.write(pdf)

        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export data as Excel (data only, no formatting)

        Query params: period, branch, start_date, end_date

        Returns: Excel file with raw data in separate sheets
        """
        period = request.query_params.get('period', 'week')
        branch_id = request.query_params.get('branch')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        start_date, end_date = self.get_date_range(period, start_date_str, end_date_str)

        # Fetch all report data
        request._request.GET = request._request.GET.copy()
        sales_data = self.sales(request).data
        expenses_data = self.expenses(request).data
        products_data = self.products(request).data
        trends_data = self.trends(request).data

        # Create Excel workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Sales Summary
        ws1 = wb.create_sheet('Penjualan')
        ws1.append(['Laporan Penjualan'])
        ws1.append([f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'])
        ws1.append([])
        ws1.append(['Metrik', 'Nilai'])
        ws1.append(['Total Pendapatan', sales_data['summary']['total_revenue']])
        ws1.append(['Total Pesanan', sales_data['summary']['total_orders']])
        ws1.append(['Rata-rata Nilai Pesanan', sales_data['summary']['avg_order_value']])
        ws1.append(['Total Pengeluaran', sales_data['summary']['total_expenses']])
        ws1.append(['Laba Bersih', sales_data['summary']['net_profit']])
        ws1.append(['Margin Laba (%)', sales_data['summary']['profit_margin']])
        ws1.append([])
        ws1.append(['Rincian Harian'])
        ws1.append(['Tanggal', 'Pendapatan', 'Pesanan', 'Rata-rata Nilai', 'Produk Terlaris'])
        for day in sales_data['daily_breakdown']:
            ws1.append([
                day['date'],
                day['revenue'],
                day['orders'],
                day['avg_order_value'],
                day['top_product']
            ])

        # Sheet 2: Expenses
        ws2 = wb.create_sheet('Pengeluaran')
        ws2.append(['Rincian Pengeluaran'])
        ws2.append([f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'])
        ws2.append([])
        ws2.append(['Kategori', 'Jumlah', 'Persentase', 'Tren', 'Perubahan (%)'])
        for item in expenses_data['by_category']:
            ws2.append([
                item['category'],
                item['amount'],
                item['percentage'],
                item['trend'],
                item['change_percentage']
            ])

        # Sheet 3: Products
        ws3 = wb.create_sheet('Produk')
        ws3.append(['Produk Terlaris'])
        ws3.append([f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'])
        ws3.append([])
        ws3.append(['Produk', 'Terjual', 'Pendapatan', 'Laba', 'Margin (%)', 'Kontribusi (%)'])
        for item in products_data['top_products']:
            ws3.append([
                item['product_name'],
                item['quantity_sold'],
                item['revenue'],
                item['profit'],
                item['profit_margin'],
                item['contribution_percentage']
            ])

        # Sheet 4: Trends
        ws4 = wb.create_sheet('Tren')
        ws4.append(['Analisis Tren'])
        ws4.append([f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'])
        ws4.append([])
        ws4.append(['Data Harian'])
        ws4.append(['Tanggal', 'Pendapatan', 'Pesanan', 'Pelanggan Unik'])
        for item in trends_data['time_series']:
            ws4.append([
                item['date'],
                item['revenue'],
                item['orders'],
                item['unique_customers']
            ])

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="data_laporan_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx"'

        return response
