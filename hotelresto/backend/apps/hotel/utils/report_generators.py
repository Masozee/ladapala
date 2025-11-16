"""
Report generation utilities for PDF and Excel formats
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFReportGenerator:
    """Generate beautiful PDF reports"""

    def __init__(self, title, period=None):
        self.title = title
        self.period = period
        self.buffer = BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30,
        )
        self.story = []
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4E61D3'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER,
        ))

        # Section heading
        self.styles.add(ParagraphStyle(
            name='SectionHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4E61D3'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))

    def add_header(self):
        """Add report header"""
        # Title
        title = Paragraph(self.title, self.styles['CustomTitle'])
        self.story.append(title)

        # Period and generation date
        if self.period:
            period_text = f"Period: {self.period}"
        else:
            period_text = ""

        generated_text = f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}"

        if period_text:
            subtitle = Paragraph(f"{period_text}<br/>{generated_text}", self.styles['CustomSubtitle'])
        else:
            subtitle = Paragraph(generated_text, self.styles['CustomSubtitle'])

        self.story.append(subtitle)
        self.story.append(Spacer(1, 0.3 * inch))

    def add_section(self, title):
        """Add a section heading"""
        heading = Paragraph(title, self.styles['SectionHeading'])
        self.story.append(heading)

    def add_paragraph(self, text):
        """Add a paragraph"""
        para = Paragraph(text, self.styles['Normal'])
        self.story.append(para)
        self.story.append(Spacer(1, 0.1 * inch))

    def add_key_metrics(self, metrics):
        """Add key metrics as a table"""
        data = []
        for key, value in metrics.items():
            data.append([key, str(value)])

        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F7FA')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ]))

        self.story.append(table)
        self.story.append(Spacer(1, 0.2 * inch))

    def add_table(self, headers, data, col_widths=None):
        """Add a data table"""
        table_data = [headers] + data

        if col_widths is None:
            col_widths = [1.5 * inch] * len(headers)

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E61D3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ]))

        self.story.append(table)
        self.story.append(Spacer(1, 0.2 * inch))

    def add_spacer(self, height=0.2):
        """Add vertical space"""
        self.story.append(Spacer(1, height * inch))

    def generate(self):
        """Generate the PDF and return buffer"""
        self.doc.build(self.story)
        self.buffer.seek(0)
        return self.buffer


class ExcelReportGenerator:
    """Generate Excel reports with formatting"""

    def __init__(self, title, period=None):
        self.title = title
        self.period = period
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.current_sheet = None
        self.current_row = 1

        # Define styles
        self.header_fill = PatternFill(start_color='4E61D3', end_color='4E61D3', fill_type='solid')
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.title_font = Font(name='Arial', size=16, bold=True, color='4E61D3')
        self.section_font = Font(name='Arial', size=12, bold=True, color='4E61D3')
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)

        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_sheet(self, sheet_name):
        """Add a new worksheet"""
        self.current_sheet = self.workbook.create_sheet(title=sheet_name)
        self.current_row = 1

    def add_header(self):
        """Add report header"""
        sheet = self.current_sheet

        # Title
        sheet.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = sheet[f'A{self.current_row}']
        cell.value = self.title
        cell.font = self.title_font
        cell.alignment = self.center_alignment
        self.current_row += 1

        # Period and date
        if self.period:
            sheet.merge_cells(f'A{self.current_row}:E{self.current_row}')
            cell = sheet[f'A{self.current_row}']
            cell.value = f"Period: {self.period}"
            cell.alignment = self.center_alignment
            self.current_row += 1

        sheet.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = sheet[f'A{self.current_row}']
        cell.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
        cell.alignment = self.center_alignment
        self.current_row += 2  # Add spacing

    def add_section(self, title):
        """Add section heading"""
        sheet = self.current_sheet
        sheet.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = sheet[f'A{self.current_row}']
        cell.value = title
        cell.font = self.section_font
        cell.alignment = self.left_alignment
        self.current_row += 1

    def add_key_metrics(self, metrics):
        """Add key metrics"""
        sheet = self.current_sheet

        for key, value in metrics.items():
            # Key column
            cell_key = sheet[f'A{self.current_row}']
            cell_key.value = key
            cell_key.font = self.normal_font
            cell_key.alignment = self.left_alignment
            cell_key.border = self.thin_border

            # Value column
            cell_value = sheet[f'B{self.current_row}']
            cell_value.value = value
            cell_value.font = self.bold_font
            cell_value.alignment = self.right_alignment
            cell_value.border = self.thin_border

            # Fill background
            cell_key.fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
            cell_value.fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')

            self.current_row += 1

        self.current_row += 1  # Add spacing

    def add_table(self, headers, data):
        """Add data table"""
        sheet = self.current_sheet

        # Headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        self.current_row += 1

        # Data rows
        for row_data in data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=self.current_row, column=col_idx)
                cell.value = value
                cell.font = self.normal_font
                cell.alignment = self.left_alignment
                cell.border = self.thin_border

                # Alternating row colors
                if self.current_row % 2 == 0:
                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

            self.current_row += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 20

        self.current_row += 1  # Add spacing

    def generate(self):
        """Generate Excel file and return buffer"""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


def format_currency(amount):
    """Format number as Indonesian Rupiah"""
    if isinstance(amount, (int, float, Decimal)):
        return f"Rp {amount:,.0f}"
    return str(amount)


def format_percentage(value):
    """Format number as percentage"""
    if isinstance(value, (int, float, Decimal)):
        return f"{value:.1f}%"
    return str(value)
